"""Questrade XLSX activity importer.

Reads Questrade "Account Activity" XLSX exports (Reports > Activity in the
Questrade portal) and emits Beancount transactions.

Handles fully:
  - Buy / Sell trades (with cost basis, commissions)
  - Dividends (DIV action, and "None" action with Activity Type = Dividends)
  - Interest (INT)
  - Contributions / Deposits (CON) — offset to Equity:Transfers (pairs with TD)
  - FX conversions (FXT) — pairs the two same-day rows into one txn

Punted to FIXME for human review (flagged with `!`):
  - Corporate actions (DIS stock split, CIL cash-in-lieu, REV reversal)
  - GIC trades (BUY/RDM with CUSIP symbol)

Usage in importers_config.yml:
    questrade:
      importer: questrade
      account: Assets:Broker:Questrade:ACC
      account_number: "1234567890"
"""

from __future__ import annotations

import re
import warnings
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from beancount.core import amount, data, position
from beancount.core.number import D
from beangulp import Importer

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Questrade uses CUSIP-style codes for some securities; map back to tickers.
SYMBOL_MAP = {
    ".CASH": "CASH",
    "H027305": "EEMV",
    "H027336": "ACWV",
    "H031006": "IEFA",
    "V002565": "VTI",
    "W017516": "GLDM",
    "W011694": "GLDM",
    "G008158": "GME",
    "C006280": "COST",
    "A603109": "AAPL",
    "T043489": "TKO",
    "TDB353": "TDB3533",
}

INCOME_DIVIDENDS = "Income:Dividends:Questrade"
INCOME_INTEREST = "Income:Interest:Questrade"
INCOME_SEC_LENDING = "Income:SecuritiesLending:Questrade"
INCOME_CAPITAL_GAINS = "Income:CapitalGains:Questrade"
COMMISSIONS = "Expenses:Broker:Questrade:Commissions"
FIXME = "Expenses:FIXME"


def contributions_account(asset_account):
    """Derive the per-account contributions equity bucket.

    Assets:Broker:Questrade:Acc -> Equity:Contributions:Questrade:Acc
    """
    return asset_account.replace("Assets:Broker:", "Equity:Contributions:", 1)


def normalize_symbol(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    if s in SYMBOL_MAP:
        return SYMBOL_MAP[s]
    if s.endswith(".TO"):
        return s[:-3]
    return s


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value).split(" ")[0], "%Y-%m-%d").date()


def to_decimal(value):
    if value is None or value == "":
        return D("0")
    return D(str(value))


class QuestradeImporter(Importer):
    def __init__(self, account, account_number):
        self.account_arg = account
        self.account_number = str(account_number)

    def name(self):
        return f"questrade.{self.account_arg}"

    def identify(self, filepath):
        path = Path(filepath)
        if not path.name.endswith(".xlsx"):
            return False
        if not re.search(
            r"(TFSA|LIRRSP|LIRA|RRSP|RRIF|LIF|RESP|FHSA|Margin|Cash)_Activities",
            path.name,
            re.IGNORECASE,
        ):
            return False
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if row and len(row) > 11 and str(row[11] or "").strip() == self.account_number:
                    return True
                if i > 5:
                    break
        except Exception:
            return False
        return False

    def filename(self, filepath):
        return f"questrade-{Path(filepath).name}"

    def account(self, filepath):
        return self.account_arg

    def date(self, filepath):
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active
        latest = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            d = parse_date(row[0])
            if latest is None or d > latest:
                latest = d
        return latest or date.today()

    def extract(self, filepath, existing=None):
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            if str(row[11] or "").strip() != self.account_number:
                continue
            rows.append({
                "txn_date": parse_date(row[0]),
                "settle_date": parse_date(row[1]) if row[1] else parse_date(row[0]),
                "action": (row[2] or "").upper(),
                "symbol": row[3],
                "description": row[4] or "",
                "quantity": to_decimal(row[5]),
                "price": to_decimal(row[6]),
                "gross": to_decimal(row[7]),
                "commission": to_decimal(row[8]),
                "net": to_decimal(row[9]),
                "currency": row[10],
                "activity_type": row[12],
            })

        # Sort by date so the resulting beancount file is chronological
        rows.sort(key=lambda r: r["txn_date"])

        # Pair FXT rows by date (they always come in CAD/USD pairs on same date)
        entries = []
        fxt_buffer = {}
        for i, r in enumerate(rows):
            if r["action"] == "FXT":
                key = (r["txn_date"], r["description"])
                fxt_buffer.setdefault(key, []).append((i, r))
                if len(fxt_buffer[key]) == 2:
                    entries.append(self._fxt(filepath, fxt_buffer.pop(key)))
            else:
                entries.append(self._row_to_entry(filepath, r, i))

        # Any unpaired FXT becomes a FIXME
        for unpaired in fxt_buffer.values():
            for i, r in unpaired:
                entries.append(self._fixme(filepath, r, i))

        return [e for e in entries if e is not None]

    def _meta(self, filepath, lineno):
        return data.new_metadata(str(filepath), lineno)

    def _row_to_entry(self, filepath, r, lineno):
        action = r["action"]
        sym = normalize_symbol(r["symbol"])
        ccy = r["currency"]
        net = r["net"]

        # Cash-only events route by action regardless of symbol shape
        if action == "INT":
            return self._interest(filepath, r, ccy, net, lineno)
        if action == "LFJ":
            return self._sec_lending(filepath, r, ccy, net, lineno)
        if action == "CON":
            return self._contribution(filepath, r, ccy, net, lineno)
        if action == "DIV" or (action == "" and r["activity_type"] == "Dividends"):
            return self._dividend(filepath, r, sym, ccy, net, lineno)

        # Inventory-affecting events: punt unmapped 7-char alphanumeric codes
        # (Questrade-internal codes like H027305, GIC CUSIPs like 5VXBYR9) to
        # FIXME. TD mutual fund codes (TDB####) are real symbols and pass through.
        if (
            sym
            and re.match(r"^[0-9A-Z]{7}$", sym)
            and not re.match(r"^TDB\d+$", sym)
            and sym not in SYMBOL_MAP.values()
        ):
            return self._fixme(filepath, r, lineno)

        if action == "BUY" or action == "DRI":
            return self._buy(filepath, r, sym, ccy, lineno)
        if action == "SELL":
            return self._sell(filepath, r, sym, ccy, lineno)
        if action == "TF6":
            return self._transfer_in(filepath, r, sym, ccy, lineno)

        # DIS / CIL / REV / RDM / unknown → FIXME
        return self._fixme(filepath, r, lineno)

    def _buy(self, filepath, r, sym, ccy, lineno):
        units = amount.Amount(abs(r["quantity"]), sym)
        cost = position.CostSpec(r["price"], None, ccy, r["txn_date"], None, False)
        postings = [
            data.Posting(self.account_arg, units, cost, None, None, None),
            data.Posting(self.account_arg, amount.Amount(r["net"], ccy), None, None, None, None),
        ]
        if r["commission"] != 0:
            postings.append(
                data.Posting(COMMISSIONS, amount.Amount(-r["commission"], ccy),
                             None, None, None, None)
            )
        return data.Transaction(
            self._meta(filepath, lineno), r["txn_date"], "*",
            "Questrade", f"BUY {abs(r['quantity'])} {sym}",
            data.EMPTY_SET, data.EMPTY_SET, postings,
        )

    def _sell(self, filepath, r, sym, ccy, lineno):
        # The Income:CapitalGains posting (no amount) auto-balances after
        # FIFO booking — Beancount fills in the actual gain (cost vs price).
        # beancount-import's UI may show a tiny rounding residual (because
        # it computes weight at price, not cost); classify that residual to
        # Income:CapitalGains:Questrade and it merges with the auto-balance.
        units = amount.Amount(r["quantity"], sym)
        price = amount.Amount(r["price"], ccy)
        cost = position.CostSpec(None, None, None, None, None, False)
        postings = [
            data.Posting(self.account_arg, units, cost, price, None, None),
            data.Posting(self.account_arg, amount.Amount(r["net"], ccy), None, None, None, None),
            data.Posting(INCOME_CAPITAL_GAINS, None, None, None, None, None),
        ]
        if r["commission"] != 0:
            postings.append(
                data.Posting(COMMISSIONS, amount.Amount(-r["commission"], ccy),
                             None, None, None, None)
            )
        return data.Transaction(
            self._meta(filepath, lineno), r["txn_date"], "*",
            "Questrade", f"SELL {abs(r['quantity'])} {sym}",
            data.EMPTY_SET, data.EMPTY_SET, postings,
        )

    def _dividend(self, filepath, r, sym, ccy, net, lineno):
        income_acct = (
            f"{INCOME_DIVIDENDS}:{sym}" if sym and sym != "CASH" else INCOME_DIVIDENDS
        )
        narration = f"Dividend {sym}" if sym else "Dividend"
        postings = [
            data.Posting(self.account_arg, amount.Amount(net, ccy), None, None, None, None),
            data.Posting(income_acct, amount.Amount(-net, ccy), None, None, None, None),
        ]
        return data.Transaction(
            self._meta(filepath, lineno), r["txn_date"], "*",
            "Questrade", narration,
            data.EMPTY_SET, data.EMPTY_SET, postings,
        )

    def _interest(self, filepath, r, ccy, net, lineno):
        postings = [
            data.Posting(self.account_arg, amount.Amount(net, ccy), None, None, None, None),
            data.Posting(INCOME_INTEREST, amount.Amount(-net, ccy), None, None, None, None),
        ]
        return data.Transaction(
            self._meta(filepath, lineno), r["txn_date"], "*",
            "Questrade", "Interest",
            data.EMPTY_SET, data.EMPTY_SET, postings,
        )

    def _sec_lending(self, filepath, r, ccy, net, lineno):
        postings = [
            data.Posting(self.account_arg, amount.Amount(net, ccy), None, None, None, None),
            data.Posting(INCOME_SEC_LENDING, amount.Amount(-net, ccy), None, None, None, None),
        ]
        return data.Transaction(
            self._meta(filepath, lineno), r["txn_date"], "*",
            "Questrade", r["description"][:160],
            data.EMPTY_SET, data.EMPTY_SET, postings,
        )

    def _transfer_in(self, filepath, r, sym, ccy, lineno):
        # External transfer of units (e.g., "TF6" with a TDB#### symbol).
        # Establish lots at the transfer price; offset to the same equity
        # contributions bucket used by CON, since these are external assets
        # being funded into the brokerage account.
        qty = abs(r["quantity"])
        price = r["price"]
        units = amount.Amount(qty, sym)
        cost = position.CostSpec(price, None, ccy, r["txn_date"], None, False)
        total = qty * price
        postings = [
            data.Posting(self.account_arg, units, cost, None, None, None),
            data.Posting(contributions_account(self.account_arg),
                         amount.Amount(-total, ccy), None, None, None, None),
        ]
        return data.Transaction(
            self._meta(filepath, lineno), r["txn_date"], "*",
            "Questrade", f"TRANSFER-IN {qty} {sym}",
            data.EMPTY_SET, data.EMPTY_SET, postings,
        )

    def _contribution(self, filepath, r, ccy, net, lineno):
        # Offset to a per-account equity bucket that accumulates over time.
        # Bank-side TD entries are not paired here; this account represents
        # "external money funded into the brokerage account."
        postings = [
            data.Posting(self.account_arg, amount.Amount(net, ccy), None, None, None, None),
            data.Posting(contributions_account(self.account_arg),
                         amount.Amount(-net, ccy), None, None, None, None),
        ]
        return data.Transaction(
            self._meta(filepath, lineno), r["txn_date"], "*",
            "Questrade", f"Contribution {r['description']}",
            data.EMPTY_SET, data.EMPTY_SET, postings,
        )

    def _fxt(self, filepath, pair):
        # Order pair so the negative leg comes first; attach price to the
        # positive leg so the transaction balances. 10-digit precision keeps
        # rounding residual ~1e-10 (well below tolerance) without tripping
        # beancount-import's matcher on 28-digit Decimals.
        (i_a, a), (i_b, b) = sorted(pair, key=lambda x: x[1]["net"])
        rate = abs(a["net"]) / abs(b["net"]) if b["net"] != 0 else D("1")
        rate_amount = amount.Amount(rate.quantize(D("0.0000000001")), a["currency"])
        postings = [
            data.Posting(self.account_arg, amount.Amount(a["net"], a["currency"]),
                         None, None, None, None),
            data.Posting(self.account_arg, amount.Amount(b["net"], b["currency"]),
                         None, rate_amount, None, None),
        ]
        return data.Transaction(
            self._meta(filepath, i_a), a["txn_date"], "*",
            "Questrade", a["description"],
            data.EMPTY_SET, data.EMPTY_SET, postings,
        )

    def _fixme(self, filepath, r, lineno):
        ccy = r["currency"] or "CAD"
        net = r["net"]
        # Don't prefix with "FIXME" — the `!` flag already marks this for
        # review, and a stable narration keeps source_desc matching across
        # re-extracts even after the user edits the entry.
        narration = f"{r['action']} {r['symbol'] or ''} {r['description']}".strip()[:200]
        postings = [
            data.Posting(self.account_arg, amount.Amount(net, ccy), None, None, None, None),
            data.Posting(FIXME, amount.Amount(-net, ccy), None, None, None, None),
        ]
        return data.Transaction(
            self._meta(filepath, lineno), r["txn_date"], "!",
            "Questrade", narration,
            data.EMPTY_SET, data.EMPTY_SET, postings,
        )


def get_importer(account, account_number=None, **_params):
    return QuestradeImporter(account=account, account_number=account_number)


if __name__ == "__main__":
    import beangulp
    importer = get_importer("Assets:Broker:Questrade:Acc", account_number="1234567890")
    ingest = beangulp.Ingest([importer], [])
    ingest()
